"""
Audit Service
Core service for logging and querying audit events
"""

from datetime import datetime, timedelta
from typing import Optional, Dict, Any, List, Union
from io import BytesIO
import json
import csv
import logging

from app.models.audit_store import audit_db
from app.utils.datetime_utils import utc_now
from .event_types import (
    EventType, EventCategory, Severity, Action,
    get_event_type,
    ENTITY_CREATED, ENTITY_UPDATED, ENTITY_DELETED
)

logger = logging.getLogger(__name__)


class AuditService:
    """Service for logging and querying audit events"""

    def __init__(self, organization_id: str = None):
        self.organization_id = organization_id

    def log(
        self,
        event_type: Union[str, EventType],
        action: Union[str, Action],
        organization_id: str = None,
        entity_type: str = None,
        entity_id: str = None,
        entity_name: str = None,
        user_id: str = None,
        user_email: str = None,
        user_role: str = None,
        changes: Dict[str, Any] = None,
        metadata: Dict[str, Any] = None,
        ip_address: str = None,
        user_agent: str = None,
        session_id: str = None,
        request_id: str = None,
        severity: Union[str, Severity] = None,
        compliance_tags: List[str] = None,
        before_snapshot: Dict = None,
        after_snapshot: Dict = None
    ) -> Optional[Dict]:
        """
        Log an audit event

        Args:
            event_type: Event type name or EventType object
            action: Action performed
            organization_id: Organization ID (uses instance default if not provided)
            entity_type: Type of entity affected
            entity_id: ID of entity affected
            entity_name: Human-readable entity name
            user_id: ID of user performing action
            user_email: Email of user
            user_role: Role of user
            changes: Field changes {field: {old: ..., new: ...}}
            metadata: Additional metadata
            ip_address: Client IP address
            user_agent: Client user agent
            session_id: Session ID
            request_id: Request ID
            severity: Override default severity
            compliance_tags: Override default compliance tags
            before_snapshot: Complete entity state before change
            after_snapshot: Complete entity state after change

        Returns:
            Created audit log entry
        """
        try:
            org_id = organization_id or self.organization_id
            if not org_id:
                logger.warning("Audit log without organization_id")
                return None

            # Resolve event type
            if isinstance(event_type, str):
                event_type_obj = get_event_type(event_type)
                event_type_name = event_type
            else:
                event_type_obj = event_type
                event_type_name = event_type.name

            # Resolve action
            if isinstance(action, Action):
                action_str = action.value
            else:
                action_str = action

            # Resolve severity
            if severity is None:
                severity_str = event_type_obj.default_severity.value if event_type_obj else 'info'
            elif isinstance(severity, Severity):
                severity_str = severity.value
            else:
                severity_str = severity

            # Resolve compliance tags
            if compliance_tags is None and event_type_obj:
                compliance_tags = event_type_obj.compliance_tags or []

            # Build log entry data
            log_data = {
                "organization_id": org_id,
                "event_type": event_type_name,
                "event_category": event_type_obj.category.value if event_type_obj else "system",
                "severity": severity_str,
                "entity_type": entity_type,
                "entity_id": entity_id,
                "entity_name": entity_name,
                "user_id": user_id,
                "user_email": user_email,
                "user_role": user_role,
                "ip_address": ip_address,
                "user_agent": user_agent,
                "session_id": session_id,
                "request_id": request_id,
                "action": action_str,
                "changes": changes,
                "metadata": metadata or {},
                "compliance_tags": compliance_tags,
            }

            # Create audit log
            audit_log = audit_db.audit_logs.create(log_data)

            # Create change history if snapshots provided
            if before_snapshot is not None or after_snapshot is not None:
                changed_fields = list(changes.keys()) if changes else []

                audit_db.change_history.create({
                    "audit_log_id": audit_log["id"],
                    "organization_id": org_id,
                    "entity_type": entity_type,
                    "entity_id": entity_id,
                    "before_snapshot": before_snapshot,
                    "after_snapshot": after_snapshot,
                    "changed_fields": changed_fields,
                    "change_summary": self._generate_change_summary(changes),
                })

            logger.debug(f"Audit logged: {event_type_name} on {entity_type}/{entity_id}")
            return audit_log

        except Exception as e:
            logger.error(f"Failed to log audit event: {e}")
            return None

    def log_create(
        self,
        entity_type: str,
        entity_id: str,
        entity_name: str = None,
        data: Dict = None,
        **kwargs
    ) -> Optional[Dict]:
        """Log entity creation"""
        return self.log(
            event_type=ENTITY_CREATED,
            action=Action.CREATE,
            entity_type=entity_type,
            entity_id=entity_id,
            entity_name=entity_name,
            after_snapshot=data,
            **kwargs
        )

    def log_update(
        self,
        entity_type: str,
        entity_id: str,
        entity_name: str = None,
        changes: Dict[str, Dict] = None,
        before_snapshot: Dict = None,
        after_snapshot: Dict = None,
        **kwargs
    ) -> Optional[Dict]:
        """Log entity update"""
        return self.log(
            event_type=ENTITY_UPDATED,
            action=Action.UPDATE,
            entity_type=entity_type,
            entity_id=entity_id,
            entity_name=entity_name,
            changes=changes,
            before_snapshot=before_snapshot,
            after_snapshot=after_snapshot,
            **kwargs
        )

    def log_delete(
        self,
        entity_type: str,
        entity_id: str,
        entity_name: str = None,
        before_snapshot: Dict = None,
        **kwargs
    ) -> Optional[Dict]:
        """Log entity deletion"""
        return self.log(
            event_type=ENTITY_DELETED,
            action=Action.DELETE,
            entity_type=entity_type,
            entity_id=entity_id,
            entity_name=entity_name,
            before_snapshot=before_snapshot,
            severity=Severity.WARNING,
            **kwargs
        )

    def _generate_change_summary(self, changes: Dict) -> str:
        """Generate human-readable change summary"""
        if not changes:
            return ""

        parts = []
        for field, change in list(changes.items())[:5]:
            old_val = change.get('old', 'null')
            new_val = change.get('new', 'null')

            # Truncate long values
            if isinstance(old_val, str) and len(old_val) > 50:
                old_val = old_val[:50] + '...'
            if isinstance(new_val, str) and len(new_val) > 50:
                new_val = new_val[:50] + '...'

            parts.append(f"{field}: {old_val} -> {new_val}")

        return "; ".join(parts)

    # Query methods

    def get_logs(
        self,
        filters: Dict = None,
        page: int = 1,
        per_page: int = 50
    ) -> tuple:
        """Get audit logs with pagination"""
        org_filters = filters.copy() if filters else {}
        org_filters["organization_id"] = self.organization_id

        offset = (page - 1) * per_page
        logs = audit_db.audit_logs.find_all(org_filters, limit=per_page, offset=offset)
        total = audit_db.audit_logs.count(org_filters)

        return logs, {
            "page": page,
            "per_page": per_page,
            "total": total,
            "pages": (total + per_page - 1) // per_page
        }

    def get_log(self, log_id: str) -> Optional[Dict]:
        """Get audit log by ID"""
        log = audit_db.audit_logs.find_by_id(log_id)
        if log and log.get("organization_id") == self.organization_id:
            return log
        return None

    def get_entity_audit_trail(
        self,
        entity_type: str,
        entity_id: str,
        limit: int = 50
    ) -> List[Dict]:
        """Get audit trail for an entity"""
        return audit_db.audit_logs.find_all({
            "organization_id": self.organization_id,
            "entity_type": entity_type,
            "entity_id": entity_id
        }, limit=limit)

    def get_user_activity(
        self,
        user_id: str,
        limit: int = 50
    ) -> List[Dict]:
        """Get activity for a user"""
        return audit_db.audit_logs.find_all({
            "organization_id": self.organization_id,
            "user_id": user_id
        }, limit=limit)

    def get_statistics(self, since: datetime = None) -> Dict:
        """Get audit log statistics"""
        if not since:
            since = utc_now() - timedelta(days=30)

        logs = audit_db.audit_logs.find_all({
            "organization_id": self.organization_id,
            "from_date": since.isoformat()
        }, limit=100000)

        stats = {
            "total_events": len(logs),
            "by_action": {},
            "by_entity_type": {},
            "by_user": {},
            "by_severity": {},
            "by_category": {},
        }

        for log in logs:
            # By action
            action = log.get("action", "unknown")
            stats["by_action"][action] = stats["by_action"].get(action, 0) + 1

            # By entity type
            entity_type = log.get("entity_type")
            if entity_type:
                stats["by_entity_type"][entity_type] = stats["by_entity_type"].get(entity_type, 0) + 1

            # By user
            user_email = log.get("user_email")
            if user_email:
                stats["by_user"][user_email] = stats["by_user"].get(user_email, 0) + 1

            # By severity
            severity = log.get("severity", "info")
            stats["by_severity"][severity] = stats["by_severity"].get(severity, 0) + 1

            # By category
            category = log.get("event_category", "system")
            stats["by_category"][category] = stats["by_category"].get(category, 0) + 1

        return stats

    # Export methods

    def export_to_csv(self, filters: Dict = None) -> BytesIO:
        """Export audit logs to CSV"""
        org_filters = filters.copy() if filters else {}
        org_filters["organization_id"] = self.organization_id
        logs = audit_db.audit_logs.find_all(org_filters, limit=10000)

        output = BytesIO()

        # Write BOM for Excel compatibility
        output.write(b'\xef\xbb\xbf')

        # Create CSV
        import io
        text_output = io.StringIO()
        writer = csv.writer(text_output)

        # Header
        writer.writerow([
            'ID', 'Event Type', 'Category', 'Severity', 'Action',
            'Entity Type', 'Entity ID', 'Entity Name',
            'User Email', 'User Role', 'IP Address',
            'Occurred At', 'Compliance Tags'
        ])

        # Data
        for log in logs:
            writer.writerow([
                log.get('id'),
                log.get('event_type'),
                log.get('event_category'),
                log.get('severity'),
                log.get('action'),
                log.get('entity_type'),
                log.get('entity_id'),
                log.get('entity_name'),
                log.get('user_email'),
                log.get('user_role'),
                log.get('ip_address'),
                log.get('occurred_at'),
                ','.join(log.get('compliance_tags', [])),
            ])

        output.write(text_output.getvalue().encode('utf-8'))
        output.seek(0)
        return output

    def export_to_json(self, filters: Dict = None) -> BytesIO:
        """Export audit logs to JSON"""
        org_filters = filters.copy() if filters else {}
        org_filters["organization_id"] = self.organization_id
        logs = audit_db.audit_logs.find_all(org_filters, limit=10000)

        output = BytesIO()
        output.write(json.dumps({
            "exported_at": utc_now().isoformat(),
            "organization_id": self.organization_id,
            "total_logs": len(logs),
            "logs": logs
        }, indent=2, default=str).encode('utf-8'))
        output.seek(0)
        return output

    def export_to_excel(self, filters: Dict = None) -> BytesIO:
        """Export audit logs to Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            # Fallback to CSV if openpyxl not available
            return self.export_to_csv(filters)

        org_filters = filters.copy() if filters else {}
        org_filters["organization_id"] = self.organization_id
        logs = audit_db.audit_logs.find_all(org_filters, limit=10000)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Audit Logs"

        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        # Headers
        headers = [
            'ID', 'Event Type', 'Category', 'Severity', 'Action',
            'Entity Type', 'Entity ID', 'Entity Name',
            'User Email', 'User Role', 'IP Address',
            'Occurred At', 'Compliance Tags'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Data
        for row, log in enumerate(logs, 2):
            ws.cell(row=row, column=1, value=log.get('id'))
            ws.cell(row=row, column=2, value=log.get('event_type'))
            ws.cell(row=row, column=3, value=log.get('event_category'))
            ws.cell(row=row, column=4, value=log.get('severity'))
            ws.cell(row=row, column=5, value=log.get('action'))
            ws.cell(row=row, column=6, value=log.get('entity_type'))
            ws.cell(row=row, column=7, value=log.get('entity_id'))
            ws.cell(row=row, column=8, value=log.get('entity_name'))
            ws.cell(row=row, column=9, value=log.get('user_email'))
            ws.cell(row=row, column=10, value=log.get('user_role'))
            ws.cell(row=row, column=11, value=log.get('ip_address'))
            ws.cell(row=row, column=12, value=log.get('occurred_at'))
            ws.cell(row=row, column=13, value=','.join(log.get('compliance_tags', [])))

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # Integrity methods

    def verify_integrity(self) -> tuple:
        """Verify audit log chain integrity"""
        return audit_db.audit_logs.verify_chain(self.organization_id)

    def get_chain_status(self) -> Dict:
        """Get chain status summary"""
        return audit_db.audit_logs.get_chain_status(self.organization_id)


# Global service instance (for use without organization context)
audit_service = AuditService()


# Convenience functions
def log_audit(event_type: str, action: str, organization_id: str = None, **kwargs) -> Optional[Dict]:
    """Log audit event using global service"""
    service = AuditService(organization_id)
    return service.log(event_type=event_type, action=action, **kwargs)


def log_create(entity_type: str, entity_id: str, organization_id: str = None, **kwargs) -> Optional[Dict]:
    """Log entity creation"""
    service = AuditService(organization_id)
    return service.log_create(entity_type=entity_type, entity_id=entity_id, **kwargs)


def log_update(entity_type: str, entity_id: str, organization_id: str = None, **kwargs) -> Optional[Dict]:
    """Log entity update"""
    service = AuditService(organization_id)
    return service.log_update(entity_type=entity_type, entity_id=entity_id, **kwargs)


def log_delete(entity_type: str, entity_id: str, organization_id: str = None, **kwargs) -> Optional[Dict]:
    """Log entity deletion"""
    service = AuditService(organization_id)
    return service.log_delete(entity_type=entity_type, entity_id=entity_id, **kwargs)
