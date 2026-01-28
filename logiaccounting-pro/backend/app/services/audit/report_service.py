"""
Report Service
Generate compliance and audit reports in multiple formats
"""

from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional
from io import BytesIO
import json
import logging

from app.models.audit_store import audit_db
from app.utils.datetime_utils import utc_now
from .compliance_service import ComplianceService

logger = logging.getLogger(__name__)


class ReportService:
    """Service for generating audit and compliance reports"""

    def __init__(self, organization_id: str):
        self.organization_id = organization_id

    def generate_compliance_report(
        self,
        framework: str,
        include_evidence: bool = True
    ) -> Dict[str, Any]:
        """
        Generate compliance report for a framework

        Args:
            framework: Framework ID (sox, gdpr, soc2)
            include_evidence: Include evidence details

        Returns:
            Report data
        """
        compliance_service = ComplianceService(self.organization_id)
        result = compliance_service.run_framework_checks(framework)

        report = {
            'report_type': 'compliance_summary',
            'framework': framework,
            'organization_id': self.organization_id,
            'generated_at': utc_now().isoformat(),
            'summary': result['summary'],
            'controls': result['controls'],
        }

        if not include_evidence:
            for control in report['controls']:
                control.pop('evidence', None)

        return report

    def generate_activity_report(
        self,
        start_date: datetime,
        end_date: datetime,
        user_id: str = None,
        entity_type: str = None
    ) -> Dict[str, Any]:
        """Generate activity summary report"""
        filters = {
            "organization_id": self.organization_id,
            "from_date": start_date.isoformat(),
            "to_date": end_date.isoformat()
        }

        if user_id:
            filters["user_id"] = user_id
        if entity_type:
            filters["entity_type"] = entity_type

        logs = audit_db.audit_logs.find_all(filters, limit=10000)

        # Aggregate statistics
        stats = {
            'total_events': len(logs),
            'by_action': {},
            'by_entity_type': {},
            'by_user': {},
            'by_severity': {},
        }

        for log in logs:
            # By action
            action = log.get("action", "unknown")
            stats['by_action'][action] = stats['by_action'].get(action, 0) + 1

            # By entity type
            etype = log.get("entity_type")
            if etype:
                stats['by_entity_type'][etype] = stats['by_entity_type'].get(etype, 0) + 1

            # By user
            user_email = log.get("user_email")
            if user_email:
                stats['by_user'][user_email] = stats['by_user'].get(user_email, 0) + 1

            # By severity
            severity = log.get("severity", "info")
            stats['by_severity'][severity] = stats['by_severity'].get(severity, 0) + 1

        return {
            'report_type': 'activity_summary',
            'organization_id': self.organization_id,
            'period': {
                'start': start_date.isoformat(),
                'end': end_date.isoformat(),
            },
            'filters': {
                'user_id': user_id,
                'entity_type': entity_type,
            },
            'generated_at': utc_now().isoformat(),
            'statistics': stats,
            'events': logs[:1000],  # Limit events
        }

    def generate_access_report(
        self,
        start_date: datetime,
        end_date: datetime,
        include_failed: bool = True
    ) -> Dict[str, Any]:
        """Generate access/authentication report"""
        filters = {
            "organization_id": self.organization_id,
            "from_date": start_date.isoformat(),
            "to_date": end_date.isoformat()
        }

        if not include_failed:
            filters["success"] = True

        logs = audit_db.access_logs.find_all(filters, limit=10000)

        # Statistics
        stats = {
            'total_events': len(logs),
            'successful_logins': len([l for l in logs if l.get("success") and l.get("event_type") == "login_success"]),
            'failed_logins': len([l for l in logs if not l.get("success") and l.get("event_type") == "login_failed"]),
            'unique_users': len(set(l.get("user_email") for l in logs if l.get("user_email"))),
            'unique_ips': len(set(l.get("ip_address") for l in logs if l.get("ip_address"))),
            'by_auth_method': {},
            'by_event_type': {},
            'high_risk_events': 0,
        }

        for log in logs:
            auth_method = log.get("auth_method")
            if auth_method:
                stats['by_auth_method'][auth_method] = stats['by_auth_method'].get(auth_method, 0) + 1

            event_type = log.get("event_type")
            stats['by_event_type'][event_type] = stats['by_event_type'].get(event_type, 0) + 1

            risk_score = log.get("risk_score")
            if risk_score and risk_score >= 70:
                stats['high_risk_events'] += 1

        return {
            'report_type': 'access_review',
            'organization_id': self.organization_id,
            'period': {
                'start': start_date.isoformat(),
                'end': end_date.isoformat(),
            },
            'generated_at': utc_now().isoformat(),
            'statistics': stats,
            'events': logs[:500],
        }

    def generate_change_report(
        self,
        entity_type: str,
        entity_id: str = None,
        start_date: datetime = None,
        end_date: datetime = None
    ) -> Dict[str, Any]:
        """Generate entity change report"""
        if entity_id:
            changes = audit_db.change_history.find_entity_history(entity_type, entity_id, limit=500)
        else:
            # Get all changes for entity type from audit logs
            filters = {
                "organization_id": self.organization_id,
                "entity_type": entity_type
            }
            if start_date:
                filters["from_date"] = start_date.isoformat()
            if end_date:
                filters["to_date"] = end_date.isoformat()

            logs = audit_db.audit_logs.find_all(filters, limit=500)

            # Get associated change history
            changes = []
            for log in logs:
                if log.get("entity_id"):
                    history = audit_db.change_history.find_entity_history(
                        entity_type, log.get("entity_id"), limit=5
                    )
                    changes.extend(history)

            # Deduplicate
            seen = set()
            unique_changes = []
            for c in changes:
                if c["id"] not in seen:
                    seen.add(c["id"])
                    unique_changes.append(c)
            changes = unique_changes[:500]

        return {
            'report_type': 'change_report',
            'organization_id': self.organization_id,
            'entity_type': entity_type,
            'entity_id': entity_id,
            'period': {
                'start': start_date.isoformat() if start_date else None,
                'end': end_date.isoformat() if end_date else None,
            },
            'generated_at': utc_now().isoformat(),
            'total_changes': len(changes),
            'changes': changes,
        }

    def generate_pdf_report(self, report_data: Dict[str, Any]) -> BytesIO:
        """Generate PDF from report data"""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        except ImportError:
            # Fallback to text if reportlab not available
            return self._generate_text_report(report_data)

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []

        # Title
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
        )

        report_type = report_data.get('report_type', 'Report').replace('_', ' ').title()
        elements.append(Paragraph(report_type, title_style))

        # Metadata
        elements.append(Paragraph(f"Generated: {report_data.get('generated_at', '')}", styles['Normal']))
        elements.append(Paragraph(f"Organization: {report_data.get('organization_id', '')}", styles['Normal']))
        elements.append(Spacer(1, 20))

        # Summary section
        if 'summary' in report_data:
            elements.append(Paragraph("Summary", styles['Heading2']))
            summary = report_data['summary']

            summary_data = [
                ['Metric', 'Value'],
                ['Overall Score', f"{summary.get('overall_score', 0):.1f}%"],
                ['Status', summary.get('status', 'Unknown')],
                ['Controls Passed', str(summary.get('passed', 0))],
                ['Controls Failed', str(summary.get('failed', 0))],
                ['Warnings', str(summary.get('warnings', 0))],
            ]

            table = Table(summary_data, colWidths=[200, 200])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))

            elements.append(table)
            elements.append(Spacer(1, 20))

        # Statistics section
        if 'statistics' in report_data:
            elements.append(Paragraph("Statistics", styles['Heading2']))
            stats = report_data['statistics']

            for key, value in stats.items():
                if not isinstance(value, dict):
                    elements.append(Paragraph(f"{key.replace('_', ' ').title()}: {value}", styles['Normal']))

            elements.append(Spacer(1, 20))

        doc.build(elements)
        buffer.seek(0)

        return buffer

    def _generate_text_report(self, report_data: Dict[str, Any]) -> BytesIO:
        """Generate text report as fallback"""
        buffer = BytesIO()

        lines = [
            f"Report: {report_data.get('report_type', 'Unknown')}",
            f"Generated: {report_data.get('generated_at', '')}",
            f"Organization: {report_data.get('organization_id', '')}",
            "",
        ]

        if 'summary' in report_data:
            lines.append("SUMMARY")
            lines.append("-" * 40)
            summary = report_data['summary']
            lines.append(f"Overall Score: {summary.get('overall_score', 0):.1f}%")
            lines.append(f"Status: {summary.get('status', 'Unknown')}")
            lines.append(f"Passed: {summary.get('passed', 0)}")
            lines.append(f"Failed: {summary.get('failed', 0)}")
            lines.append("")

        if 'statistics' in report_data:
            lines.append("STATISTICS")
            lines.append("-" * 40)
            for key, value in report_data['statistics'].items():
                if not isinstance(value, dict):
                    lines.append(f"{key}: {value}")
            lines.append("")

        buffer.write("\n".join(lines).encode('utf-8'))
        buffer.seek(0)
        return buffer

    def generate_excel_report(self, report_data: Dict[str, Any]) -> BytesIO:
        """Generate Excel from report data"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            # Fallback to JSON if openpyxl not available
            return self._generate_json_report(report_data)

        buffer = BytesIO()
        wb = openpyxl.Workbook()

        # Summary sheet
        ws = wb.active
        ws.title = "Summary"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        row = 1
        ws.cell(row=row, column=1, value="Report Type").font = header_font
        ws.cell(row=row, column=1).fill = header_fill
        ws.cell(row=row, column=2, value=report_data.get('report_type', ''))
        row += 1

        ws.cell(row=row, column=1, value="Generated At").font = header_font
        ws.cell(row=row, column=1).fill = header_fill
        ws.cell(row=row, column=2, value=report_data.get('generated_at', ''))
        row += 2

        # Write summary if exists
        if 'summary' in report_data:
            ws.cell(row=row, column=1, value="Summary").font = Font(bold=True)
            row += 1

            for key, value in report_data['summary'].items():
                ws.cell(row=row, column=1, value=key.replace('_', ' ').title())
                ws.cell(row=row, column=2, value=str(value))
                row += 1

        # Controls sheet
        if 'controls' in report_data:
            controls_sheet = wb.create_sheet("Controls")

            headers = ['Control ID', 'Name', 'Status', 'Score', 'Findings']
            for col, header in enumerate(headers, 1):
                cell = controls_sheet.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill

            for row_num, control in enumerate(report_data['controls'], start=2):
                controls_sheet.cell(row=row_num, column=1, value=control.get('control_id', ''))
                controls_sheet.cell(row=row_num, column=2, value=control.get('control_name', ''))
                controls_sheet.cell(row=row_num, column=3, value=control.get('status', ''))
                controls_sheet.cell(row=row_num, column=4, value=control.get('score', 0))
                controls_sheet.cell(row=row_num, column=5, value='; '.join(control.get('findings', [])))

        # Events sheet
        if 'events' in report_data and report_data['events']:
            events_sheet = wb.create_sheet("Events")

            headers = list(report_data['events'][0].keys())[:10]  # Limit columns
            for col, header in enumerate(headers, 1):
                cell = events_sheet.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill

            for row_num, event in enumerate(report_data['events'][:1000], start=2):
                for col, key in enumerate(headers, 1):
                    events_sheet.cell(row=row_num, column=col, value=str(event.get(key, '')))

        wb.save(buffer)
        buffer.seek(0)

        return buffer

    def _generate_json_report(self, report_data: Dict[str, Any]) -> BytesIO:
        """Generate JSON report"""
        buffer = BytesIO()
        buffer.write(json.dumps(report_data, indent=2, default=str).encode('utf-8'))
        buffer.seek(0)
        return buffer

    # Report types list

    @staticmethod
    def get_available_reports() -> List[Dict]:
        """Get list of available report types"""
        return [
            {
                'id': 'compliance_summary',
                'name': 'Compliance Summary',
                'description': 'Overview of compliance status across frameworks',
            },
            {
                'id': 'activity_summary',
                'name': 'Activity Summary',
                'description': 'Summary of system activity',
            },
            {
                'id': 'access_review',
                'name': 'Access Review',
                'description': 'Authentication and access events',
            },
            {
                'id': 'change_report',
                'name': 'Change Report',
                'description': 'Entity changes over time',
            },
        ]
