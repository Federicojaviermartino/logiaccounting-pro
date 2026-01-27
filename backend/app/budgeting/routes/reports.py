"""API routes for budget reports."""
from typing import Optional
from uuid import UUID
from io import BytesIO

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.database import get_db
from app.auth.dependencies import get_current_user
from app.auth.models import User
from app.budgeting.services.budget_service import BudgetService
from app.budgeting.services.variance_service import VarianceService

router = APIRouter(prefix="/reports", tags=["Budget Reports"])


@router.get("/budgets/{budget_id}/summary")
async def get_budget_summary_report(
    budget_id: UUID,
    format: str = Query("json", pattern="^(json|xlsx|pdf)$"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get budget summary report."""
    service = BudgetService(db, current_user.customer_id)
    budget = await service.get_budget_by_id(budget_id, include_versions=True)

    if format == "json":
        return {
            "budget": {
                "id": str(budget.id),
                "code": budget.budget_code,
                "name": budget.name,
                "fiscal_year": budget.fiscal_year,
                "status": budget.status.value,
                "total_revenue": float(budget.total_revenue),
                "total_expenses": float(budget.total_expenses),
                "total_net_income": float(budget.total_net_income),
            },
            "versions": [
                {
                    "id": str(v.id),
                    "name": v.version_name,
                    "number": v.version_number,
                    "status": v.status.value,
                    "is_active": v.is_active,
                }
                for v in budget.versions
            ]
        }

    elif format == "xlsx":
        # Generate Excel report
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Budget Summary"

        # Header
        ws['A1'] = "Budget Summary Report"
        ws['A2'] = f"Budget: {budget.name} ({budget.budget_code})"
        ws['A3'] = f"Fiscal Year: {budget.fiscal_year}"

        # Summary
        ws['A5'] = "Category"
        ws['B5'] = "Amount"
        ws['A6'] = "Total Revenue"
        ws['B6'] = float(budget.total_revenue)
        ws['A7'] = "Total Expenses"
        ws['B7'] = float(budget.total_expenses)
        ws['A8'] = "Net Income"
        ws['B8'] = float(budget.total_net_income)

        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=budget_{budget.budget_code}.xlsx"}
        )

    elif format == "pdf":
        # Return simple text for now - would use reportlab in production
        return {"message": "PDF generation not implemented yet"}


@router.get("/budgets/{budget_id}/variance")
async def get_variance_report(
    budget_id: UUID,
    period_type: str = Query("ytd", pattern="^(monthly|quarterly|ytd|annual)$"),
    year: Optional[int] = None,
    month: Optional[int] = Query(None, ge=1, le=12),
    format: str = Query("json", pattern="^(json|xlsx)$"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get variance analysis report."""
    variance_service = VarianceService(db, current_user.customer_id)
    report = await variance_service.get_budget_vs_actual(budget_id, period_type, year, month)

    if format == "json":
        return report

    elif format == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Variance Analysis"

        # Header
        ws['A1'] = "Variance Analysis Report"
        ws['A2'] = f"Budget: {report.budget_name}"
        ws['A3'] = f"Period: {report.period_type.upper()} {report.period_year}"

        # Column headers
        headers = ["Account Code", "Account Name", "Budget", "Actual", "Variance", "Variance %", "Status"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=5, column=col, value=header)
            ws.cell(row=5, column=col).font = Font(bold=True)

        # Revenue section
        ws['A7'] = "REVENUE"
        ws['A7'].font = Font(bold=True)
        row = 8
        for line in report.revenue_lines:
            ws.cell(row=row, column=1, value=line.account_code)
            ws.cell(row=row, column=2, value=line.account_name)
            ws.cell(row=row, column=3, value=float(line.budgeted))
            ws.cell(row=row, column=4, value=float(line.actual))
            ws.cell(row=row, column=5, value=float(line.variance))
            ws.cell(row=row, column=6, value=f"{float(line.variance_percent):.1f}%")
            ws.cell(row=row, column=7, value=line.variance_type)
            row += 1

        # Expense section
        row += 1
        ws.cell(row=row, column=1, value="EXPENSES")
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        for line in report.expense_lines:
            ws.cell(row=row, column=1, value=line.account_code)
            ws.cell(row=row, column=2, value=line.account_name)
            ws.cell(row=row, column=3, value=float(line.budgeted))
            ws.cell(row=row, column=4, value=float(line.actual))
            ws.cell(row=row, column=5, value=float(line.variance))
            ws.cell(row=row, column=6, value=f"{float(line.variance_percent):.1f}%")
            ws.cell(row=row, column=7, value=line.variance_type)
            row += 1

        # Totals
        row += 1
        ws.cell(row=row, column=1, value="NET INCOME")
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=3, value=float(report.net_income_budget))
        ws.cell(row=row, column=4, value=float(report.net_income_actual))
        ws.cell(row=row, column=5, value=float(report.net_income_variance))

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=variance_report_{budget_id}.xlsx"}
        )


@router.get("/budgets/{budget_id}/export")
async def export_budget(
    budget_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Export full budget to Excel."""
    from openpyxl import Workbook

    service = BudgetService(db, current_user.customer_id)
    budget = await service.get_budget_by_id(budget_id, include_versions=True)

    wb = Workbook()

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary['A1'] = budget.name
    ws_summary['A2'] = f"Code: {budget.budget_code}"
    ws_summary['A3'] = f"Fiscal Year: {budget.fiscal_year}"
    ws_summary['A4'] = f"Period: {budget.start_date} to {budget.end_date}"
    ws_summary['A5'] = f"Status: {budget.status.value}"
    ws_summary['A7'] = "Total Revenue"
    ws_summary['B7'] = float(budget.total_revenue)
    ws_summary['A8'] = "Total Expenses"
    ws_summary['B8'] = float(budget.total_expenses)
    ws_summary['A9'] = "Net Income"
    ws_summary['B9'] = float(budget.total_net_income)

    # Version sheets
    for version in budget.versions:
        if version.is_active:
            ws = wb.create_sheet(title=f"v{version.version_number}_{version.version_name[:20]}")

            # Headers
            headers = ["Account Code", "Account Name", "Type", "Annual",
                       "Jan", "Feb", "Mar", "Apr", "May", "Jun",
                       "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

            # This would need to fetch lines - simplified for now
            ws['A2'] = "Budget lines would go here..."

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={budget.budget_code}_export.xlsx"}
    )
