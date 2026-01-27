"""Excel report generator."""
from io import BytesIO
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill

from app.reporting.schemas.financial_statements import (
    BalanceSheetData, IncomeStatementData, TrialBalanceData
)


class ExcelReportGenerator:
    """Generate Excel reports for financial statements."""
    
    def __init__(self):
        self.header_font = Font(bold=True, size=14)
        self.title_font = Font(bold=True, size=12)
        self.bold_font = Font(bold=True)
        self.currency_format = '#,##0.00'
        self.thin_border = Border(bottom=Side(style='thin'))
        self.double_border = Border(bottom=Side(style='double'))
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font_white = Font(bold=True, color="FFFFFF")
    
    def generate_balance_sheet(self, data: BalanceSheetData) -> BytesIO:
        """Generate Balance Sheet Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Balance Sheet"
        
        ws.merge_cells('A1:D1')
        ws['A1'] = data.company_name
        ws['A1'].font = self.header_font
        
        ws.merge_cells('A2:D2')
        ws['A2'] = f"Balance Sheet as of {data.report_date.strftime('%B %d, %Y')}"
        ws['A2'].font = self.title_font
        
        row = 4
        
        ws[f'A{row}'] = "Account"
        ws[f'B{row}'] = "Current Period"
        for col in ['A', 'B']:
            ws[f'{col}{row}'].font = self.header_font_white
            ws[f'{col}{row}'].fill = self.header_fill
        
        row += 2
        
        ws[f'A{row}'] = "ASSETS"
        ws[f'A{row}'].font = self.bold_font
        row += 1
        
        ws[f'A{row}'] = "Current Assets"
        ws[f'A{row}'].font = self.bold_font
        row += 1
        
        for line in data.current_assets:
            ws[f'A{row}'] = "  " * line.indent_level + line.label
            ws[f'B{row}'] = float(line.current_period)
            ws[f'B{row}'].number_format = self.currency_format
            row += 1
        
        ws[f'A{row}'] = "Total Current Assets"
        ws[f'A{row}'].font = self.bold_font
        ws[f'B{row}'] = float(data.total_current_assets)
        ws[f'B{row}'].number_format = self.currency_format
        ws[f'B{row}'].font = self.bold_font
        ws[f'B{row}'].border = self.thin_border
        row += 2
        
        ws[f'A{row}'] = "Non-Current Assets"
        ws[f'A{row}'].font = self.bold_font
        row += 1
        
        for line in data.non_current_assets:
            ws[f'A{row}'] = "  " * line.indent_level + line.label
            ws[f'B{row}'] = float(line.current_period)
            ws[f'B{row}'].number_format = self.currency_format
            row += 1
        
        ws[f'A{row}'] = "Total Non-Current Assets"
        ws[f'A{row}'].font = self.bold_font
        ws[f'B{row}'] = float(data.total_non_current_assets)
        ws[f'B{row}'].number_format = self.currency_format
        ws[f'B{row}'].font = self.bold_font
        ws[f'B{row}'].border = self.thin_border
        row += 2
        
        ws[f'A{row}'] = "TOTAL ASSETS"
        ws[f'A{row}'].font = self.bold_font
        ws[f'B{row}'] = float(data.total_assets)
        ws[f'B{row}'].number_format = self.currency_format
        ws[f'B{row}'].font = self.bold_font
        ws[f'B{row}'].border = self.double_border
        row += 3
        
        ws[f'A{row}'] = "LIABILITIES"
        ws[f'A{row}'].font = self.bold_font
        row += 1
        
        ws[f'A{row}'] = "Current Liabilities"
        ws[f'A{row}'].font = self.bold_font
        row += 1
        
        for line in data.current_liabilities:
            ws[f'A{row}'] = "  " * line.indent_level + line.label
            ws[f'B{row}'] = float(line.current_period)
            ws[f'B{row}'].number_format = self.currency_format
            row += 1
        
        ws[f'A{row}'] = "Total Current Liabilities"
        ws[f'A{row}'].font = self.bold_font
        ws[f'B{row}'] = float(data.total_current_liabilities)
        ws[f'B{row}'].number_format = self.currency_format
        ws[f'B{row}'].font = self.bold_font
        ws[f'B{row}'].border = self.thin_border
        row += 2
        
        ws[f'A{row}'] = "Non-Current Liabilities"
        ws[f'A{row}'].font = self.bold_font
        row += 1
        
        for line in data.non_current_liabilities:
            ws[f'A{row}'] = "  " * line.indent_level + line.label
            ws[f'B{row}'] = float(line.current_period)
            ws[f'B{row}'].number_format = self.currency_format
            row += 1
        
        ws[f'A{row}'] = "Total Non-Current Liabilities"
        ws[f'A{row}'].font = self.bold_font
        ws[f'B{row}'] = float(data.total_non_current_liabilities)
        ws[f'B{row}'].number_format = self.currency_format
        ws[f'B{row}'].font = self.bold_font
        ws[f'B{row}'].border = self.thin_border
        row += 2
        
        ws[f'A{row}'] = "TOTAL LIABILITIES"
        ws[f'A{row}'].font = self.bold_font
        ws[f'B{row}'] = float(data.total_liabilities)
        ws[f'B{row}'].number_format = self.currency_format
        ws[f'B{row}'].font = self.bold_font
        ws[f'B{row}'].border = self.thin_border
        row += 3
        
        ws[f'A{row}'] = "EQUITY"
        ws[f'A{row}'].font = self.bold_font
        row += 1
        
        for line in data.equity:
            ws[f'A{row}'] = "  " * line.indent_level + line.label
            ws[f'B{row}'] = float(line.current_period)
            ws[f'B{row}'].number_format = self.currency_format
            row += 1
        
        ws[f'A{row}'] = "TOTAL EQUITY"
        ws[f'A{row}'].font = self.bold_font
        ws[f'B{row}'] = float(data.total_equity)
        ws[f'B{row}'].number_format = self.currency_format
        ws[f'B{row}'].font = self.bold_font
        ws[f'B{row}'].border = self.thin_border
        row += 2
        
        ws[f'A{row}'] = "TOTAL LIABILITIES & EQUITY"
        ws[f'A{row}'].font = self.bold_font
        ws[f'B{row}'] = float(data.total_liabilities_and_equity)
        ws[f'B{row}'].number_format = self.currency_format
        ws[f'B{row}'].font = self.bold_font
        ws[f'B{row}'].border = self.double_border
        
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 18
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def generate_income_statement(self, data: IncomeStatementData) -> BytesIO:
        """Generate Income Statement Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Income Statement"
        
        ws.merge_cells('A1:C1')
        ws['A1'] = data.company_name
        ws['A1'].font = self.header_font
        
        ws.merge_cells('A2:C2')
        ws['A2'] = f"Income Statement: {data.period_start.strftime('%b %d, %Y')} - {data.period_end.strftime('%b %d, %Y')}"
        ws['A2'].font = self.title_font
        
        row = 4
        
        ws[f'A{row}'] = "Account"
        ws[f'B{row}'] = "Amount"
        ws[f'C{row}'] = "% of Revenue"
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = self.header_font_white
            ws[f'{col}{row}'].fill = self.header_fill
        
        row += 2
        
        ws[f'A{row}'] = "REVENUE"
        ws[f'A{row}'].font = self.bold_font
        row += 1
        
        for line in data.revenue:
            ws[f'A{row}'] = "  " + line.label
            ws[f'B{row}'] = float(line.current_period)
            ws[f'B{row}'].number_format = self.currency_format
            row += 1
        
        ws[f'A{row}'] = "Total Revenue"
        ws[f'A{row}'].font = self.bold_font
        ws[f'B{row}'] = float(data.total_revenue)
        ws[f'B{row}'].number_format = self.currency_format
        ws[f'B{row}'].font = self.bold_font
        ws[f'B{row}'].border = self.thin_border
        ws[f'C{row}'] = "100.00%"
        row += 2
        
        ws[f'A{row}'] = "COST OF SALES"
        ws[f'A{row}'].font = self.bold_font
        row += 1
        
        for line in data.cost_of_sales:
            ws[f'A{row}'] = "  " + line.label
            ws[f'B{row}'] = float(line.current_period)
            ws[f'B{row}'].number_format = self.currency_format
            row += 1
        
        ws[f'A{row}'] = "Total Cost of Sales"
        ws[f'A{row}'].font = self.bold_font
        ws[f'B{row}'] = float(data.total_cost_of_sales)
        ws[f'B{row}'].number_format = self.currency_format
        ws[f'B{row}'].font = self.bold_font
        ws[f'B{row}'].border = self.thin_border
        row += 2
        
        ws[f'A{row}'] = "GROSS PROFIT"
        ws[f'A{row}'].font = self.bold_font
        ws[f'B{row}'] = float(data.gross_profit)
        ws[f'B{row}'].number_format = self.currency_format
        ws[f'B{row}'].font = self.bold_font
        if data.gross_margin_percent:
            ws[f'C{row}'] = f"{float(data.gross_margin_percent):.2f}%"
        row += 2
        
        ws[f'A{row}'] = "OPERATING EXPENSES"
        ws[f'A{row}'].font = self.bold_font
        row += 1
        
        for line in data.operating_expenses:
            ws[f'A{row}'] = "  " + line.label
            ws[f'B{row}'] = float(line.current_period)
            ws[f'B{row}'].number_format = self.currency_format
            row += 1
        
        ws[f'A{row}'] = "Total Operating Expenses"
        ws[f'A{row}'].font = self.bold_font
        ws[f'B{row}'] = float(data.total_operating_expenses)
        ws[f'B{row}'].number_format = self.currency_format
        ws[f'B{row}'].font = self.bold_font
        ws[f'B{row}'].border = self.thin_border
        row += 2
        
        ws[f'A{row}'] = "NET INCOME"
        ws[f'A{row}'].font = self.bold_font
        ws[f'B{row}'] = float(data.net_income)
        ws[f'B{row}'].number_format = self.currency_format
        ws[f'B{row}'].font = self.bold_font
        ws[f'B{row}'].border = self.double_border
        if data.net_margin_percent:
            ws[f'C{row}'] = f"{float(data.net_margin_percent):.2f}%"
        
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def generate_trial_balance(self, data: TrialBalanceData) -> BytesIO:
        """Generate Trial Balance Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Trial Balance"
        
        ws.merge_cells('A1:D1')
        ws['A1'] = data.company_name
        ws['A1'].font = self.header_font
        
        ws.merge_cells('A2:D2')
        ws['A2'] = f"Trial Balance as of {data.as_of_date.strftime('%B %d, %Y')}"
        ws['A2'].font = self.title_font
        
        row = 4
        
        headers = ['Account Code', 'Account Name', 'Debit', 'Credit']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font_white
            cell.fill = self.header_fill
        
        row += 1
        
        for line in data.lines:
            ws.cell(row=row, column=1, value=line.account_code)
            ws.cell(row=row, column=2, value=line.account_name)
            
            debit_cell = ws.cell(row=row, column=3, value=float(line.debit) if line.debit else None)
            debit_cell.number_format = self.currency_format
            
            credit_cell = ws.cell(row=row, column=4, value=float(line.credit) if line.credit else None)
            credit_cell.number_format = self.currency_format
            
            row += 1
        
        row += 1
        ws.cell(row=row, column=2, value="TOTALS").font = self.bold_font
        
        total_debit = ws.cell(row=row, column=3, value=float(data.total_debits))
        total_debit.number_format = self.currency_format
        total_debit.font = self.bold_font
        total_debit.border = self.double_border
        
        total_credit = ws.cell(row=row, column=4, value=float(data.total_credits))
        total_credit.number_format = self.currency_format
        total_credit.font = self.bold_font
        total_credit.border = self.double_border
        
        row += 2
        status = "BALANCED" if data.is_balanced else f"OUT OF BALANCE: {data.difference}"
        ws.cell(row=row, column=2, value=status).font = self.bold_font
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
